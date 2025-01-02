import sqlite3
from datetime import datetime

import openpyxl
from dateutil.relativedelta import relativedelta


def del_tuple_in_list(data: list) -> list:
    """
    将形如[('1',), ('2',), ('3',),]的数据转化为[1, 2, 3,]
    :param data:带有元组的列表
    :return: 清洗后的列表
    """

    output = []

    output.extend(single_data[0] for single_data in data)

    return output


def execute_sql_sentence(sentence: str, ) -> list:
    """
    执行数据库语句并返回列表
    :param sentence: 需要执行的语句
    :return:
    """

    c, conn = connect_database()

    result = []
    try:
        c.execute(sentence)
        result = c.fetchall()

    except Exception as e:
        print(e)

    disconnect_database(conn=conn)

    return result


# kind:"在编","编外"
def connect_database() -> tuple[sqlite3.Cursor, sqlite3.Connection]:
    """
    用于连接数据库
    :return:
    """

    conn = sqlite3.connect(
        fr"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\database\educational_data.db"
    )
    c = conn.cursor()

    return c, conn


def disconnect_database(conn) -> None:
    """
    用于断开数据库
    :param conn:
    :return:
    """

    conn.close()

    return None


def read_xlsx_to_list(file_path, sheet_name=None):
    # 打开工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 如果未指定工作表名称，则使用第一个工作表
    if sheet_name is None:
        sheet = workbook.active
    else:
        sheet = workbook[sheet_name]

    # 读取数据并存储到二维列表中
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data


def save_excel(two_dimension_list: list[list or tuple], excel_name: str = "output"):
    # 读取 Excel 文件
    workbook = openpyxl.Workbook()

    # 获取活动的工作表
    ws = workbook.active

    for item in two_dimension_list:
        ws.append(item)

    workbook.save(f"{excel_name}.xlsx")


def get_age_from_citizen_id(citizen_id: str, year: str = None, month: int = 9, day: int = 1) -> int:
    """
    通过身份证号计算当前年龄或截止某一年某一月某一日（默认某一年的9月1日）\n
    get_age_from_citizen_id(citizen_id = "440105200102220000") -> 23
    :param citizen_id: 身份证号
    :param year: 截止年份
    :param month: 截止月份
    :param day: 截止日期
    :return: 年龄,两位数int
    """

    if len(citizen_id) != 18:
        return -1

    try:
        if year is None:
            return max(
                relativedelta(
                    dt1=datetime.today(),
                    dt2=datetime(
                        year=int(citizen_id[6:10]),
                        month=int(citizen_id[10:12]),
                        day=int(citizen_id[12:14])
                    )
                ).years,
                0
            )

        elif 2000 <= int(year) <= 3000:
            return max(
                relativedelta(
                    dt1=datetime(year=int(year), month=month, day=day),
                    dt2=datetime(
                        year=int(citizen_id[6:10]),
                        month=int(citizen_id[10:12]),
                        day=int(citizen_id[12:14])
                    )

                ).years,
                0
            )

        else:
            return -2

    except Exception as e:
        print(text=f"{e}:{citizen_id}")
        return -3


def get_educational_background_order() -> dict:
    """
    学历排序
    :return:
    """

    return {
        '博士研究生': 1, '硕士研究生': 2, '本科': 3, "专科": 4, "高中": 5, "高中及以下": 6,
        "中师": 7, "中专（非师范）": 8, "中专": 9, "初中": 10, None: 11
    }


def get_period_order() -> dict:
    """
    学段排序
    :return:
    """

    return {'高中': 1, '初中': 2, '小学': 3, '幼儿园': 4, '中职': 5, "其他": 6, None: 7,}
