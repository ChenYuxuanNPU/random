import os
import shutil

from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from func import *

db_path = r"C:\Users\1012986131\Desktop\python\streamlit_pyecharts\database\educational_data.db"

title_1 = [
    "单位全称（按照单位公章）", "学校类型", "统一社会信用代码", "姓名", "身份证号码（文本格式）", "性别",
    "民族", "籍贯（精确到市）", "婚姻状况", "政治面貌", "入党（团）时间", "参加工作前学历",
    "参加工作前学位", "参加工作前毕业院校（学信网全称）", "参加工作前毕业时间（文本格式xxxx年xx月，其他日期同）",
    "参加工作前所学专业全称", "最高学历",
    "最高学位", "最高学历毕业院校学信网全称",
    "最高学历毕业时间", "最高学历所学专业全称", "参加工作时间", "首次进入白云教育时间", "进入现单位任教时间",
    "任现行政职务级别（如正校级、副校级、中层正副职）", "任现行政职务级别的时间",
    "任教年级", "主教学科（只能选一门）", "现持有最高职称", "职称证专业",
    "最高职称证书认定或评审通过时间", "普通话层次", "教师资格层次", "教师资格学科",
    "最高骨干教师级别", "四名工作室主持人名称（最高级别）",
    "四名工作室主持人确定时间（最高级别）", "曾获市级及以上综合荣誉（可填多个，逗号隔开）",
    "曾获市级以上人才项目称号（可填多个，逗号隔开）",
    "现常住地址", "手机号码", "紧急联系人", "与联系人关系", "紧急联系人手机号码", "备注", "区域",
    "任教学段", "是否音体美专任教师", "年龄", "教师身份"
]

list_b = ["十二年一贯制", "完全中学", "高中", "中职", "九年一贯制", "初中", "小学", "幼儿园", "教学支撑单位"]
list_f = ["男", "女"]
list_i = ["未婚", "已婚", "离异", "丧偶", "复婚", "再婚"]
list_j = ["群众", "中共党员", "共青团员", "民进会员", "民盟盟员", "致公党员", "中共预备党员", "民革会员",
          "九三学社", "农工党员", "无党派人士", "其他"]
list_l = ["博士研究生", "硕士研究生", "本科", "专科", "中专（非师范）", "中师", "高中", "初中"]
list_m = ["博士学位", "硕士学位", "学士学位", "无"]
list_q = ["博士研究生", "硕士研究生", "本科", "专科", "中专（非师范）", "中师", "高中", "初中"]
list_r = ["博士学位", "硕士学位", "学士学位", "无"]
list_y = ["党组织书记", "党组织书记兼校长", "正校级", "副校级", "中层正职", "中层副职", "少先队大队辅导员",
          "少先队副大队辅导员", "工会主席", "工会副主席", "团委书记", "团委副书记", "无"]
list_ab = ["语文", "数学", "英语", "思想政治", "历史", "地理", "物理", "化学", "生物", "体育", "音乐", "美术",
           "书法", "舞蹈", "科学", "信息技术", "通用技术", "劳动", "综合实践", "心理健康", "人工智能", "汽修",
           "烹饪", "幼儿教育", "电子商务", "特殊教育", "校本课程", "其他", "无"]
list_ac = ["正高级教师", "高级教师", "一级教师", "二级教师", "三级教师", "高级职称（非中小学系列）",
           "中级职称（非中小学系列）", "初级职称（非中小学系列）", "未取得职称"]
list_af = ["一级甲等", "一级乙等", "二级甲等", "二级乙等", "三级甲等", "三级乙等", "无"]
list_ag = ["幼儿园", "小学", "初级中学", "高级中学", "中职专业课", "中职实习指导教师", "高等学校", "无"]
list_ai = ["广东省骨干教师", "广州市骨干教师", "白云区骨干教师", "其他", "无"]
list_at = ["直管", "永平", "石井", "新市", "江高", "人和", "太和", "钟落潭"]
list_au = ["高中", "初中", "小学", "中职", "幼儿园", "其他"]
list_av = ["是", "否"]
list_ax = ["政府雇员", "公办临聘教师", "民办学校教师"]


def write_list_to_txt(route: str):
    with open(file=route, mode='w', encoding='UTF-8') as f:
        for i in (["B列"] + list_b):
            f.write(i + "\n")
        f.write("\n")

        for i in (["F列"] + list_f):
            f.write(i + "\n")
        f.write("\n")

        for i in (["I列"] + list_i):
            f.write(i + "\n")
        f.write("\n")

        for i in (["J列"] + list_j):
            f.write(i + "\n")
        f.write("\n")

        for i in (["L列"] + list_l):
            f.write(i + "\n")
        f.write("\n")

        for i in (["M列"] + list_m):
            f.write(i + "\n")
        f.write("\n")

        for i in (["Q列"] + list_q):
            f.write(i + "\n")
        f.write("\n")

        for i in (["R列"] + list_r):
            f.write(i + "\n")
        f.write("\n")

        for i in (["Y列"] + list_y):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AB列"] + list_ab):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AC列"] + list_ac):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AF列"] + list_af):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AG列"] + list_ag):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AI列"] + list_ai):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AT列"] + list_at):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AU列"] + list_au):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AV列"] + list_av):
            f.write(i + "\n")
        f.write("\n")

        for i in (["AX列"] + list_ax):
            f.write(i + "\n")
        f.write("\n")


def write_name_list_to_txt(route: str, name_list: list):
    with open(file=route, mode='w', encoding='UTF-8') as f:
        for name in name_list:
            f.write(name + "\n")


def ensure_folders_exist_or_clear(kind: str, area_name1="", school_name1=""):
    folder_names = ["直管", "永平", "江高", "石井", "新市", "人和", "太和", "钟落潭"]

    if school_name1 == "":
        base_dir = fr"C:\Users\1012986131\Desktop\python\random\update_data_source\output\{kind}"

        """  
        确保在base_dir路径下存在列表中的每个文件夹，如果不存在则创建，如果存在则清空（删除并重建）。  

        :param base_dir: 文件夹的基路径  
        :param folder_names: 列表，包含要检查/创建的文件夹名称  
        """
        for folder_name in folder_names:
            full_path = os.path.join(base_dir, folder_name)

            # 检查文件夹是否存在
            if os.path.exists(full_path):
                # 如果存在，则删除并重新创建
                shutil.rmtree(full_path)  # 删除文件夹及其所有内容
                os.makedirs(full_path)  # 重新创建文件夹
            else:
                # 如果不存在，则创建文件夹
                os.makedirs(full_path)

    else:
        base_dir = fr"C:\Users\1012986131\Desktop\python\random\update_data_source\output\{kind}\{area_name1}"
        full_path = os.path.join(base_dir, school_name1)

        # 检查文件夹是否存在
        if os.path.exists(full_path):
            # 如果存在，则删除并重新创建
            shutil.rmtree(full_path)  # 删除文件夹及其所有内容
            os.makedirs(full_path)  # 重新创建文件夹
        else:
            # 如果不存在，则创建文件夹
            os.makedirs(full_path)


# 这一段是用来判断是否有空单元格的，原理是遍历每一列，检查每一列的非空单元格数，若所有列的最大值和最小值相同，则认为所有信息填写完成
def check_blank():
    # 初始化公式的基础部分
    formula = "=IF(MAX("

    # 初始化计数器，用于列名的生成
    column_letter = 'A'

    # 循环遍历从A到BS的所有列
    for i in range(1, 27):  # BS是Excel中的第702列
        # 生成列名（A-Z）
        column_name = chr(ord(column_letter) + i - 1)
        # 拼接COUNTA函数，注意添加范围（例如A1:A500）
        if i == 1:
            formula += f"COUNTA(Sheet1!{column_name}:{column_name})"
        else:
            formula += f",COUNTA(Sheet1!{column_name}:{column_name})"

    for i in range(1, 25):  # BS是Excel中的第702列
        # 生成列名（AA-AZ）
        column_name = chr(ord(column_letter) + i - 1)

        formula += f",COUNTA(Sheet1!A{column_name}:A{column_name})"

    formula += ")=MIN("

    # 循环遍历从A到BS的所有列
    for i in range(1, 27):  # BS是Excel中的第702列
        # 生成列名（A-Z）
        column_name = chr(ord(column_letter) + i - 1)
        # 拼接COUNTA函数，注意添加范围（例如A1:A500）
        if i == 1:
            formula += f"COUNTA(Sheet1!{column_name}:{column_name})"
        else:
            formula += f",COUNTA(Sheet1!{column_name}:{column_name})"

    for i in range(1, 25):  # BS是Excel中的第702列
        # 生成列名（AA-AZ）
        column_name = chr(ord(column_letter) + i - 1)

        formula += f",COUNTA(Sheet1!A{column_name}:A{column_name})"

    formula += "),1,0)"

    return formula


def restrict(options: list, chara, ws):
    # 创建一个数据验证对象，设置为列表类型，并设置允许的选项
    dv = DataValidation(type="list", formula1='"' + ','.join(map(str, options)) + '"')

    # 将数据验证添加到特定单元格（例如A1）
    for row in range(2, 1000):  # 假设我们设置到第1000行
        cell_ref = f'{chara}{row}'  # 构建单元格引用，如A1, A2, ..., A10
        # 添加到数据验证，然后将其应用到特定的单元格
        ws.add_data_validation(dv)
        dv.add(ws[cell_ref])


def auto_adjust_column_width(worksheet, column):
    column_letter = get_column_letter(column)
    length = len(str(worksheet[column_letter][0].value))
    adjusted_width = (length + 8) * 1.5
    worksheet.column_dimensions[column_letter].width = adjusted_width


def set_all_cell_borders(ws, border_style="thin", border_color="000000"):
    """
    为工作表中的所有单元格设置边框。

    :param ws: openpyxl.worksheet.worksheet.Worksheet 对象
    :param border_style: 边框的样式，如 "thin", "medium", "thick" 等
    :param border_color: 边框的颜色，格式为 "RRGGAZ"
    """
    thin = Side(border_style=border_style, color=border_color)
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 遍历工作表中的每一行和每一列
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border


# 加入下拉选择框
def add_restrict(ws):
    restrict(options=list_b, chara="B", ws=ws)
    restrict(options=list_f, chara="F", ws=ws)
    restrict(options=list_i, chara="I", ws=ws)
    restrict(options=list_j, chara="J", ws=ws)
    restrict(options=list_l, chara="L", ws=ws)
    restrict(options=list_m, chara="M", ws=ws)
    restrict(options=list_q, chara="Q", ws=ws)
    restrict(options=list_r, chara="R", ws=ws)
    restrict(options=list_y, chara="Y", ws=ws)
    restrict(options=list_ab, chara="AB", ws=ws)
    restrict(options=list_ac, chara="AC", ws=ws)
    restrict(options=list_af, chara="AF", ws=ws)
    restrict(options=list_ag, chara="AG", ws=ws)
    restrict(options=list_ai, chara="AI", ws=ws)
    restrict(options=list_at, chara="AT", ws=ws)
    restrict(options=list_au, chara="AU", ws=ws)
    restrict(options=list_av, chara="AV", ws=ws)
    restrict(options=list_ax, chara="AX", ws=ws)


# 加入背景色
def add_bg_colour(ws):
    # 创建一个填充样式
    fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # 应用填充样式到列A的所有单元格
    for cell in ws['B']:
        cell.fill = fill
    for cell in ws['F']:
        cell.fill = fill
    for cell in ws['I']:
        cell.fill = fill
    for cell in ws['J']:
        cell.fill = fill
    for cell in ws['L']:
        cell.fill = fill
    for cell in ws['M']:
        cell.fill = fill
    for cell in ws['Q']:
        cell.fill = fill
    for cell in ws['R']:
        cell.fill = fill
    for cell in ws['Y']:
        cell.fill = fill
    for cell in ws['AB']:
        cell.fill = fill
    for cell in ws['AC']:
        cell.fill = fill
    for cell in ws['AF']:
        cell.fill = fill
    for cell in ws['AG']:
        cell.fill = fill
    for cell in ws['AI']:
        cell.fill = fill
    for cell in ws['AT']:
        cell.fill = fill
    for cell in ws['AU']:
        cell.fill = fill
    for cell in ws['AV']:
        cell.fill = fill
    for cell in ws['AX']:
        cell.fill = fill


def set_col_width(ws):
    for i in range(len(title_1) + 10):
        auto_adjust_column_width(ws, i + 1)


# 这里是用来判断填写的内容是否符合下拉列表的
def gen_list_check_sentence(words: list, location: str):
    return f"{'+'.join([f'COUNTIF(Sheet1!{location}:{location},"{value}")' for value in words])}"


# 这里是用来判断理论上相同的一列（如校名）的总数，相当于填写的信息量
def count_same_info(location: str):
    return f"=if(COUNTIF(Sheet1!{location}:{location},Sheet1!{location}2)=COUNTA(Sheet1!{location}:{location})-1,1,0)"


# 这里用来判断位数是否统一
def check_data_length(location: str, length: int):
    return f'=IF(SUMPRODUCT(--(LEN(Sheet1!{location}:{location})={length}))=COUNTA(Sheet1!A:A)-1,1,0)'


# 这里检查日期是不是按照xxxx年xx月格式
def check_str_date(location: str):
    return f'=IF(COUNTIF(Sheet1!{location}:{location},"????年??月")+COUNTIF(Sheet1!{location}:{location},"????年?月")+COUNTIF(Sheet1!{location}:{location},"无")=COUNTA(Sheet1!A:A)-1,1,0)'


# 检查带函数的单元格是否被修改为常数
def check_formula(td_list: list):
    formula = "=IF(ISFORMULA("
    formula += td_list[0]

    for i in range(1, len(td_list)):
        formula += f")*ISFORMULA({td_list[i]}"

    formula += "),1,0)"

    return formula


def add_excel_func(ws):
    # 为检查的字体添加红色
    font = Font(color="FF0000")

    # 填写完成的判断
    ws['A1'] = '检查无误'
    ws[
        'A2'] = "=IF(AND(A5=1,A8=1,C2=1,C5=1,C8=1,C11=1,E2=1,E5=1,E8=1,E11=1,E14=1,E17=1,E20=1,E23=1,E26=1,E29=1,E32=1,E35=1,E38=1,E41=1,E44=1,E47=1,G2=1,G5=1,G8=1,G11=1,G14=1,G17=1,G20=1,G23=1,G26=1,I2=1,I5=1,I8=1),1,0)"
    ws['A2'].font = font

    ws['A4'] = '是否已填写完成'
    ws['A5'] = check_blank()
    ws['A5'].font = font

    ws['A7'] = '公式是否完整'
    ws['A8'] = check_formula(
        td_list=["A5", "C2", "C5", "C8", "C11", "E2", "E5", "E8", "E11", "E14", "E17", "E20",
                 "E23", "E26", "E29", "E32", "E35", "E38", "E41", "E44", "E47",
                 "G2", "G5", "G8", "G11", "G14", "G17", "G20", "G23", "G26", "I2", "I5", "I8"])
    ws['A8'].font = font

    ws['A10'] = '空单元格个数(需要调整AX后行序号)'
    ws['A11'] = "=COUNTBLANK(Sheet1!A2:Sheet1!AX100)"
    ws['A11'].font = font

    # print(gen_list_check_sentence(words=["博士学位", "硕士学位", "学士学位", "无"],location="Y"))
    # 首先判断前三列学校信息会不会有不同的
    ws['C1'] = 'A列是否相同'
    ws['C2'] = count_same_info(location="A")
    ws['C2'].font = font
    ws['C4'] = 'B列是否相同'
    ws['C5'] = count_same_info(location="B")
    ws['C5'].font = font
    ws['C7'] = 'C列是否相同'
    ws['C8'] = count_same_info(location="C")
    ws['C8'].font = font
    ws['C10'] = 'AT列是否相同'
    ws['C11'] = count_same_info(location="AT")
    ws['C11'].font = font

    # 然后判断下拉列表的内容是否被更改
    ws['E1'] = 'B列下拉列表'
    ws['E2'] = f'=if(({gen_list_check_sentence(words=list_b, location="B")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E2'].font = font

    ws['E4'] = 'F列下拉列表'
    ws['E5'] = f'=if(({gen_list_check_sentence(words=list_f, location="F")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E5'].font = font

    ws['E7'] = 'I列下拉列表'
    ws['E8'] = f'=if(({gen_list_check_sentence(words=list_i, location="I")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E8'].font = font

    ws['E10'] = 'J列下拉列表'
    ws['E11'] = f'=if(({gen_list_check_sentence(words=list_j, location="J")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E11'].font = font

    ws['E13'] = 'L列下拉列表'
    ws['E14'] = f'=if(({gen_list_check_sentence(words=list_l, location="L")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E14'].font = font

    ws['E16'] = 'M列下拉列表'
    ws['E17'] = f'=if(({gen_list_check_sentence(words=list_m, location="M")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E17'].font = font

    ws['E19'] = 'Q列下拉列表'
    ws['E20'] = f'=if(({gen_list_check_sentence(words=list_q, location="Q")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E20'].font = font

    ws['E22'] = 'R列下拉列表'
    ws['E23'] = f'=if(({gen_list_check_sentence(words=list_r, location="R")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E23'].font = font

    ws['E25'] = 'Y列下拉列表'
    ws['E26'] = f'=if(({gen_list_check_sentence(words=list_y, location="Y")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E26'].font = font

    ws['E28'] = 'AB列下拉列表'
    ws['E29'] = f'=if(({gen_list_check_sentence(words=list_ab, location="AB")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E29'].font = font

    ws['E31'] = 'AC列下拉列表'
    ws['E32'] = f'=if(({gen_list_check_sentence(words=list_ac, location="AC")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E32'].font = font

    ws['E34'] = 'AF列下拉列表'
    ws['E35'] = f'=if(({gen_list_check_sentence(words=list_af, location="AF")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E35'].font = font

    ws['E37'] = 'AG列下拉列表'
    ws['E38'] = f'=if(({gen_list_check_sentence(words=list_ag, location="AG")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E38'].font = font

    ws['E40'] = 'AI列下拉列表'
    ws['E41'] = f'=if(({gen_list_check_sentence(words=list_ai, location="AI")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E41'].font = font

    ws['E43'] = 'AT列下拉列表'
    ws['E44'] = f'=if(({gen_list_check_sentence(words=list_at, location="AT")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E44'].font = font

    ws['E46'] = 'AU列下拉列表'
    ws['E47'] = f'=if(({gen_list_check_sentence(words=list_au, location="AU")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E47'].font = font

    ws['E46'] = 'AV列下拉列表'
    ws['E47'] = f'=if(({gen_list_check_sentence(words=list_av, location="AV")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E47'].font = font

    ws['E46'] = 'AX列下拉列表'
    ws['E47'] = f'=if(({gen_list_check_sentence(words=list_ax, location="AX")})=COUNTA(Sheet1!A:A)-1,1,0)'
    ws['E47'].font = font

    # 这里判断年月格式
    ws['G1'] = 'K列日期格式'
    ws['G2'] = check_str_date(location='K')
    ws['G2'].font = font

    ws['G4'] = 'O列日期格式'
    ws['G5'] = check_str_date(location='O')
    ws['G5'].font = font

    ws['G7'] = 'T列日期格式'
    ws['G8'] = check_str_date(location='T')
    ws['G8'].font = font

    ws['G10'] = 'V列日期格式'
    ws['G11'] = check_str_date(location='V')
    ws['G11'].font = font

    ws['G13'] = 'W列日期格式'
    ws['G14'] = check_str_date(location='W')
    ws['G14'].font = font

    ws['G16'] = 'X列日期格式'
    ws['G17'] = check_str_date(location='X')
    ws['G17'].font = font

    ws['G19'] = 'Z列日期格式'
    ws['G20'] = check_str_date(location='Z')
    ws['G20'].font = font

    ws['G22'] = 'AE列日期格式'
    ws['G23'] = check_str_date(location='AE')
    ws['G23'].font = font

    ws['G25'] = 'AK列日期格式'
    ws['G26'] = check_str_date(location='AK')
    ws['G26'].font = font

    # 这里判断身份证位数
    ws['I1'] = 'E列位数'
    ws['I2'] = check_data_length(location='E', length=18)
    ws['I2'].font = font
    # 手机号位数
    ws['I4'] = 'AO列位数'
    ws['I5'] = check_data_length(location='AO', length=11)
    ws['I5'].font = font

    ws['I7'] = 'AR列位数'
    ws['I8'] = check_data_length(location='AR', length=11)
    ws['I8'].font = font


def output_excel_1(title: list, data: list, file_name: str, area_name: str):
    # 这里是函数第一句
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(title)

    area = area_name

    name_list = []
    for data_row in data:
        name_list.append(data_row[3])

        ws.append(list(map(str, data_row)))

    add_restrict(ws=ws)

    add_bg_colour(ws=ws)

    # 开一个新表用来加入函数限制
    ws_check = wb.create_sheet(title="Sheet2")

    add_excel_func(ws=ws_check)

    set_col_width(ws=ws)
    set_col_width(ws_check)

    set_all_cell_borders(ws)
    set_all_cell_borders(ws_check)

    ensure_folders_exist_or_clear(kind="编外", school_name1=file_name, area_name1=area_name)

    write_list_to_txt(
        route=fr"C:\Users\1012986131\Desktop\python\random\update_data_source\output\编外\{area}\{file_name}\下拉列表内容.txt")

    write_name_list_to_txt(
        route=fr"C:\Users\1012986131\Desktop\python\random\update_data_source\output\编外\{area}\{file_name}\已有人员名单.txt",
        name_list=name_list)

    wb.save(
        fr"C:\Users\1012986131\Desktop\python\random\update_data_source\output\编外\{area}\{file_name}\{file_name}.xlsx")
    # print(file_name)


if __name__ == '__main__':

    result = execute_sql_sentence(
        sentence=f'select * from teacher_data_1 where "采集年份" == "2024"'
    )

    for i in range(len(result)):
        result[i] = list(result[i])
        result[i].pop(0)

    print(result)

    school_list = execute_sql_sentence(
        sentence=f'select distinct "校名" from teacher_data_1 where "采集年份" = "2024"'
    )

    print(f"编外数据共包含{len(del_tuple_in_list(school_list))}所学校")
    school_count = len(del_tuple_in_list(school_list))

    # 汇总相同学校的在编教师
    result_all = {}

    for data in result:

        if data[0] not in result_all.keys():
            result_all[data[0]] = [list(data)]

        else:
            result_all[data[0]].append(list(data))

    # print(result_all)

    ensure_folders_exist_or_clear(kind="编外")

    count = 1
    # people_count = 0
    for school, data in result_all.items():
        excel_name = school
        area_name = data[0][-5]

        # people_count += len(data)

        output_excel_1(title=title_1, data=data, file_name=excel_name, area_name=area_name)

        print("")
        print(excel_name)
        print(area_name)
        print(f"{count}/{school_count}")

        count += 1

    # print(people_count)

    # print(execute_sql_sentence(sentence=f'select count(*) from teacher_data_1 where "采集年份" == "2024"'))
