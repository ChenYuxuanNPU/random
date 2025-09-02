import openpyxl
import re

from openpyxl.utils import get_column_letter


def remove_zero_width_chars(text):
    # 移除所有零宽字符（包括\u200c, \u200d, \u200e, \u200f, \uFEFF等）
    return re.sub(r'[\u200c\u200d\u200e\u200f\uFEFF]', '', text)


def save_excel(two_dimension_list: list[list | tuple], excel_name: str = "output"):
    # 读取 Excel 文件
    workbook = openpyxl.Workbook()

    # 获取活动的工作表
    ws = workbook.active

    for item in two_dimension_list:
        ws.append(item)

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)  # 获取列字母

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2) * 1.8  # 加一些缓冲空间
        ws.column_dimensions[column_letter].width = adjusted_width

    workbook.save(f"{excel_name}.xlsx")


def read_xlsx_to_list(file_path, sheet_name="Sheet1"):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active if sheet_name is None else workbook[sheet_name]

    data = []
    for row in sheet.iter_rows(values_only=True):
        processed_row = []
        for cell_value in row:
            if isinstance(cell_value, str):
                cell_value = remove_zero_width_chars(cell_value).strip()
            processed_row.append(cell_value)
        data.append(processed_row)

    return data


output = {}

for name in ["教师问卷", "学生问卷"]:
    data = read_xlsx_to_list(file_path=f"{name}.xlsx")

    title = data[0]

    for item in data[1:]:
        if item[6] not in output.keys():
            output[item[6]] = [item]

        else:
            output[item[6]].append(item)

    for key, value in output.items():
        save_excel(two_dimension_list=[title] + value, excel_name=f"{name}/{key}")
