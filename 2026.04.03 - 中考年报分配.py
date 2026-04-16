import math


def allocate_books(data, total_books):
    """
    按学生数比例分配书籍，确保每所学校至少1本

    参数:
    data: 二维列表，格式为 [[区, 校名, 学生数], [区, 校名, 学生数], ...]
    total_books: 书籍总数 X

    返回:
    二维列表，格式为 [[区, 校名, 学生数, 分配书数], ...]
    """
    num_schools = len(data)

    # 检查书籍数量是否足够
    if total_books < num_schools:
        print(f"警告：书籍总数({total_books})少于学校数({num_schools})，无法保证每校至少1本！")
        print("将按比例分配（部分学校可能得0本）")
        guarantee_one = False
    else:
        guarantee_one = True
        print(f"确保每所学校至少获得1本书，共需{num_schools}本基础书")

    # 计算学生总数
    total_students = sum(row[2] for row in data)

    if total_students == 0:
        # 如果没有学生，平均分配
        if guarantee_one and total_books >= num_schools:
            avg_books = (total_books - num_schools) // num_schools
            result = []
            remainder = total_books - num_schools - avg_books * num_schools
            for i, row in enumerate(data):
                books = 1 + avg_books + (1 if i < remainder else 0)
                result.append([row[0], row[1], row[2], books])
        else:
            avg_books = total_books // num_schools
            result = []
            remainder = total_books - avg_books * num_schools
            for i, row in enumerate(data):
                books = avg_books + (1 if i < remainder else 0)
                result.append([row[0], row[1], row[2], books])
        return result

    # 第一步：分配基础书（每校1本）
    if guarantee_one:
        remaining_books = total_books - num_schools
        base_books = 1
    else:
        remaining_books = total_books
        base_books = 0

    # 第二步：按学生数比例分配剩余书籍
    theoretical = []
    for row in data:
        if total_students > 0:
            ratio = row[2] / total_students
            exact_books = remaining_books * ratio
        else:
            exact_books = 0
        theoretical.append([row[0], row[1], row[2], exact_books])

    # 第三步：向下取整
    allocated = []
    total_allocated = 0
    fractional_parts = []

    for i, item in enumerate(theoretical):
        full_books = math.floor(item[3])
        allocated.append([item[0], item[1], item[2], base_books + full_books])
        total_allocated += full_books
        # 记录小数部分和索引
        fractional_parts.append((item[3] - full_books, i))

    # 第四步：分配剩余书籍（按小数部分从大到小）
    remaining_after_floor = remaining_books - total_allocated
    fractional_parts.sort(reverse=True)  # 按小数部分降序排列

    for i in range(remaining_after_floor):
        if i < len(fractional_parts):
            idx = fractional_parts[i][1]
            allocated[idx][3] += 1

    # 第五步：验证每校至少1本（如果保证的话）
    if guarantee_one:
        for row in allocated:
            if row[3] < 1:
                print(f"警告：{row[1]} 分配书数为{row[3]}，但要求至少1本")

    return allocated


def allocate_books_alternative(data, total_books):
    """
    备选方案：先保证每校1本，剩余按比例分配（更直观的实现）

    参数:
    data: 二维列表，格式为 [[区, 校名, 学生数], ...]
    total_books: 书籍总数 X

    返回:
    二维列表，格式为 [[区, 校名, 学生数, 分配书数], ...]
    """
    num_schools = len(data)

    # 检查书籍数量
    if total_books < num_schools:
        print(f"错误：书籍总数({total_books})少于学校数({num_schools})，无法满足每校至少1本！")
        print("请增加书籍数量或使用无保底分配方案")
        return None

    # 计算学生总数
    total_students = sum(row[2] for row in data)

    if total_students == 0:
        # 如果没有学生，平均分配剩余书籍
        remaining_books = total_books - num_schools
        avg_extra = remaining_books // num_schools
        remainder = remaining_books - avg_extra * num_schools
        result = []
        for i, row in enumerate(data):
            books = 1 + avg_extra + (1 if i < remainder else 0)
            result.append([row[0], row[1], row[2], books])
        return result

    # 先给每所学校1本基础书
    result = []
    remaining_books = total_books - num_schools

    # 计算每所学校的理论额外分配数（基于学生数比例）
    for row in data:
        ratio = row[2] / total_students
        exact_extra = remaining_books * ratio
        result.append([row[0], row[1], row[2], 1, exact_extra])  # 临时存储：区,校名,学生数,基础1本,理论额外数

    # 向下取整
    total_extra_allocated = 0
    fractional_parts = []

    for i, item in enumerate(result):
        full_extra = math.floor(item[4])
        item[4] = full_extra  # 替换为整数部分
        total_extra_allocated += full_extra
        fractional_parts.append((item[4] - full_extra, i))  # 注意：这里小数部分应该是原始理论值的小数

        # 重新计算正确的小数部分
        original_extra = result[i][4] + (item[4] - full_extra) if i < len(result) else 0

    # 重新正确计算小数部分
    fractional_parts = []
    for i, item in enumerate(result):
        original_exact = (remaining_books * (data[i][2] / total_students))
        fractional = original_exact - math.floor(original_exact)
        fractional_parts.append((fractional, i))

    # 分配剩余书籍
    remaining_extra = remaining_books - total_extra_allocated
    fractional_parts.sort(reverse=True)

    # 构建最终结果
    final_result = []
    for i, item in enumerate(result):
        extra_books = item[4]
        final_result.append([item[0], item[1], item[2], 1 + extra_books])

    for i in range(remaining_extra):
        if i < len(fractional_parts):
            idx = fractional_parts[i][1]
            final_result[idx][3] += 1

    return final_result


def read_excel_to_list(file_path):
    """
    读取xlsx文件到二维列表（你已有的func1）
    """
    import openpyxl

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        if row[0] is not None and row[0] != '区':  # 跳过空行和表头
            # 前三列：区、校名、学生人数
            data.append([row[0], row[1], row[2]])

    return data


def write_list_to_excel(data, file_path):
    """
    将二维列表保存到xlsx文件（你已有的func2）
    """
    import openpyxl

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 添加表头
    sheet.cell(row=1, column=1, value="区")
    sheet.cell(row=1, column=2, value="校名")
    sheet.cell(row=1, column=3, value="学生人数")
    sheet.cell(row=1, column=4, value="分配书数")

    # 写入数据（从第2行开始）
    for row_idx, row in enumerate(data, 2):
        for col_idx, value in enumerate(row, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    workbook.save(file_path)


def main(input_file, output_file, total_books, guarantee_one=True):
    """
    主函数：读取学校数据，分配书籍，保存结果

    参数:
    input_file: 输入文件路径
    output_file: 输出文件路径
    total_books: 书籍总数 X
    guarantee_one: 是否保证每校至少1本（默认True）
    """
    # 1. 读取原始数据
    school_data = read_excel_to_list(input_file)
    num_schools = len(school_data)

    print(f"共有{num_schools}所学校，总书籍数：{total_books}")

    # 2. 检查书籍数量是否足够
    if guarantee_one and total_books < num_schools:
        print(f"\n错误：无法满足每校至少1本的要求！")
        print(f"需要至少{num_schools}本书，但目前只有{total_books}本")
        print("将采用无保底方案进行分配...")
        guarantee_one = False

    # 3. 分配书籍
    if guarantee_one:
        # 使用保底方案
        result_data = allocate_books(school_data, total_books)
    else:
        # 使用普通比例方案
        result_data = allocate_books(school_data, total_books)

    if result_data is None:
        print("分配失败！")
        return

    # 4. 保存结果
    write_list_to_excel(result_data, output_file)

    print(f"\n分配完成！结果已保存到 {output_file}")

    # 5. 打印分配统计
    print("\n分配结果统计：")
    print(f"{'区':<10} {'校名':<15} {'学生数':<8} {'分配书数':<8}")
    print("-" * 45)

    min_books = float('inf')
    max_books = 0
    zero_books = 0

    for row in result_data:
        print(f"{row[0]:<10} {row[1]:<15} {row[2]:<8} {row[3]:<8}")
        min_books = min(min_books, row[3])
        max_books = max(max_books, row[3])
        if row[3] == 0:
            zero_books += 1

    print(f"\n总分配书数: {sum(row[3] for row in result_data)}")
    print(f"最少获得: {min_books}本, 最多获得: {max_books}本")
    if zero_books > 0:
        print(f"警告：有{zero_books}所学校获得0本书")
    else:
        print("✓ 所有学校都至少获得1本书")

    # 6. 按区统计
    district_stats = {}
    for row in result_data:
        district = row[0]
        if district not in district_stats:
            district_stats[district] = {'students': 0, 'books': 0, 'schools': 0}
        district_stats[district]['students'] += row[2]
        district_stats[district]['books'] += row[3]
        district_stats[district]['schools'] += 1

    print("\n按区统计：")
    print(f"{'区':<10} {'学校数':<8} {'学生总数':<10} {'书总数':<8} {'人均书数':<10}")
    print("-" * 50)
    for district, stats in district_stats.items():
        avg_books_per_student = stats['books'] / stats['students'] if stats['students'] > 0 else 0
        print(
            f"{district:<10} {stats['schools']:<8} {stats['students']:<10} {stats['books']:<8} {avg_books_per_student:<10.3f}")


# 使用示例
if __name__ == "__main__":
    # 参数设置
    input_file = "schools.xlsx"  # 输入文件名（3列：区、校名、学生人数）
    output_file = "allocated_books.xlsx"  # 输出文件名（4列：区、校名、学生人数、分配书数）
    total_books = 372  # 书籍总数 X

    # 执行分配（保证每校至少1本）
    main(input_file, output_file, total_books, guarantee_one=True)